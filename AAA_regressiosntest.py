import os
import sys
import time
import json
import asyncio
import hashlib
from pathlib import Path
from typing import Dict, List, Tuple, Any

import pandas as pd
import aiohttp
from aiohttp import ClientSession, ClientTimeout, BasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from deepdiff import DeepDiff
from dotenv import load_dotenv
from auth import get_password

# =====================================================
# CONFIGURATION
# =====================================================

# Load environment variables
load_dotenv()

INPUT_FILE = "shared/reports/pl_testcases.xlsx"
OUTPUT_DIR = "shared/reports"
SOURCE_RESPONSE_DIR = "shared/reports/source_api"
TARGET_RESPONSE_DIR = "shared/reports/target_api"
REPORT_FILE = "shared/reports/report.xlsx"

# Authentication credentials
USERNAME = os.getenv("USERNAME", os.getenv("API_USERNAME", ""))
PASSWORD = None  # Will be loaded at runtime

# Performance settings
MAX_CONCURRENT_REQUESTS = 50  # Adjust based on system capacity
REQUEST_TIMEOUT = 30  # seconds
RETRY_ATTEMPTS = 3
RETRY_DELAY = 1  # seconds

# TESTING LIMIT - Process only first N test cases (for authentication testing)
# Set to None to process ALL test cases, or set to a number (e.g., 1000) to limit
LIMIT_TEST_CASES = 100  # CHANGE TO None TO PROCESS ALL TEST CASES

# =====================================================
# UTILITY FUNCTIONS
# =====================================================

def create_directories():
    """Create necessary directories if they don't exist."""
    Path(SOURCE_RESPONSE_DIR).mkdir(parents=True, exist_ok=True)
    Path(TARGET_RESPONSE_DIR).mkdir(parents=True, exist_ok=True)
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    print("[OK] Directories created successfully")


def get_safe_filename(url: str, index: int) -> str:
    """Generate a safe filename from URL."""
    # Use hash to create unique but consistent filename
    url_hash = hashlib.md5(url.encode()).hexdigest()[:8]
    return f"response_{index:05d}_{url_hash}.json"


def calculate_file_size(file_path: str) -> str:
    """Calculate file size in human-readable format."""
    if not os.path.exists(file_path):
        return "0 bytes"
    
    size_bytes = os.path.getsize(file_path)
    
    if size_bytes < 1024:
        return f"{size_bytes} bytes"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.2f} MB"


def compare_json(source_data: Any, target_data: Any) -> Tuple[bool, str]:
    """
    Compare two JSON objects and return match status and detailed mismatch info.
    
    Returns:
        Tuple[bool, str]: (is_matched, mismatch_details)
    """
    try:
        # Handle None cases
        if source_data is None and target_data is None:
            return True, "Matched"
        
        if source_data is None:
            return False, "Source response is empty/None"
        
        if target_data is None:
            return False, "Target response is empty/None"
        
        # Deep comparison
        diff = DeepDiff(source_data, target_data, ignore_order=False, verbose_level=2)
        
        if not diff:
            return True, "Matched"
        
        # Build detailed mismatch message
        mismatch_parts = []
        
        if 'values_changed' in diff:
            for key, change in list(diff['values_changed'].items())[:5]:  # Limit to first 5
                mismatch_parts.append(f"Value changed at {key}: {change['old_value']} -> {change['new_value']}")
        
        if 'dictionary_item_added' in diff:
            added = list(diff['dictionary_item_added'])[:5]
            mismatch_parts.append(f"Keys added in target: {', '.join(map(str, added))}")
        
        if 'dictionary_item_removed' in diff:
            removed = list(diff['dictionary_item_removed'])[:5]
            mismatch_parts.append(f"Keys removed from target: {', '.join(map(str, removed))}")
        
        if 'type_changes' in diff:
            for key, change in list(diff['type_changes'].items())[:3]:
                mismatch_parts.append(f"Type changed at {key}: {change['old_type']} -> {change['new_type']}")
        
        if 'iterable_item_added' in diff:
            mismatch_parts.append("Array items added in target")
        
        if 'iterable_item_removed' in diff:
            mismatch_parts.append("Array items removed from target")
        
        mismatch_message = " | ".join(mismatch_parts) if mismatch_parts else "JSON structures differ"
        
        return False, mismatch_message
    
    except Exception as e:
        return False, f"Comparison error: {str(e)}"


# =====================================================
# ASYNC API OPERATIONS
# =====================================================

class ProgressTracker:
    """Track and display progress of API operations."""
    
    def __init__(self, total: int):
        self.total = total
        self.completed = 0
        self.start_time = time.time()
        self.lock = asyncio.Lock()
    
    async def update(self, count: int = 1):
        """Update progress counter."""
        async with self.lock:
            self.completed += count
            percentage = (self.completed / self.total) * 100
            elapsed = time.time() - self.start_time
            
            # Calculate ETA
            if self.completed > 0:
                rate = self.completed / elapsed
                remaining = (self.total - self.completed) / rate if rate > 0 else 0
                eta_str = f"ETA: {int(remaining)}s"
            else:
                eta_str = "ETA: calculating..."
            
            # Print progress
            sys.stdout.write(f"\rProcessing: {percentage:.1f}% ({self.completed}/{self.total}) | {eta_str}")
            sys.stdout.flush()
    
    def finish(self):
        """Print final statistics."""
        elapsed = time.time() - self.start_time
        print(f"\n[OK] Total execution time: {elapsed:.2f} seconds")
        print(f"[OK] Average request time: {elapsed/self.total:.3f} seconds")


async def fetch_json_with_retry(
    session: ClientSession,
    url: str,
    semaphore: asyncio.Semaphore,
    progress: ProgressTracker,
    auth: BasicAuth = None
) -> Dict[str, Any]:
    """
    Fetch JSON from URL with retry logic.
    
    Returns:
        Dict with 'data', 'error', 'status_code'
    """
    async with semaphore:
        for attempt in range(RETRY_ATTEMPTS):
            try:
                async with session.get(url, auth=auth, timeout=ClientTimeout(total=REQUEST_TIMEOUT)) as response:
                    status_code = response.status
                    
                    if status_code == 200:
                        try:
                            data = await response.json()
                            await progress.update(1)
                            return {
                                'data': data,
                                'error': None,
                                'status_code': status_code
                            }
                        except json.JSONDecodeError:
                            text = await response.text()
                            await progress.update(1)
                            return {
                                'data': None,
                                'error': f"Invalid JSON response: {text[:100]}",
                                'status_code': status_code
                            }
                    else:
                        text = await response.text()
                        error_msg = f"HTTP {status_code}: {text[:100]}"
                        
                        # Don't retry on client errors (4xx)
                        if 400 <= status_code < 500:
                            await progress.update(1)
                            return {
                                'data': None,
                                'error': error_msg,
                                'status_code': status_code
                            }
                        
                        # Retry on server errors (5xx)
                        if attempt < RETRY_ATTEMPTS - 1:
                            await asyncio.sleep(RETRY_DELAY * (attempt + 1))
                            continue
                        
                        await progress.update(1)
                        return {
                            'data': None,
                            'error': error_msg,
                            'status_code': status_code
                        }
            
            except asyncio.TimeoutError:
                if attempt < RETRY_ATTEMPTS - 1:
                    await asyncio.sleep(RETRY_DELAY * (attempt + 1))
                    continue
                await progress.update(1)
                return {
                    'data': None,
                    'error': f"Timeout after {REQUEST_TIMEOUT}s",
                    'status_code': 0
                }
            
            except Exception as e:
                if attempt < RETRY_ATTEMPTS - 1:
                    await asyncio.sleep(RETRY_DELAY * (attempt + 1))
                    continue
                await progress.update(1)
                return {
                    'data': None,
                    'error': f"Request error: {str(e)}",
                    'status_code': 0
                }
        
        # Should not reach here, but just in case
        await progress.update(1)
        return {
            'data': None,
            'error': "Max retries exceeded",
            'status_code': 0
        }


async def process_test_case(
    session: ClientSession,
    semaphore: asyncio.Semaphore,
    index: int,
    source_url: str,
    target_url: str,
    progress: ProgressTracker,
    auth: BasicAuth = None
) -> Dict[str, Any]:
    """
    Process a single test case: fetch both URLs and compare.
    
    Returns:
        Dict with all test results
    """
    # Fetch source and target in parallel
    source_task = fetch_json_with_retry(session, source_url, semaphore, progress, auth)
    target_task = fetch_json_with_retry(session, target_url, semaphore, progress, auth)
    
    source_result, target_result = await asyncio.gather(source_task, target_task)
    
    # Save responses to files
    source_file = os.path.join(SOURCE_RESPONSE_DIR, get_safe_filename(source_url, index))
    target_file = os.path.join(TARGET_RESPONSE_DIR, get_safe_filename(target_url, index))
    
    # Save source response
    with open(source_file, 'w', encoding='utf-8') as f:
        json.dump(source_result, f, indent=2, ensure_ascii=False)
    
    # Save target response
    with open(target_file, 'w', encoding='utf-8') as f:
        json.dump(target_result, f, indent=2, ensure_ascii=False)
    
    # Calculate file sizes
    source_size = calculate_file_size(source_file)
    target_size = calculate_file_size(target_file)
    
    # Determine status
    if source_result['error'] or target_result['error']:
        errors = []
        if source_result['error']:
            errors.append(f"Source: {source_result['error']}")
        if target_result['error']:
            errors.append(f"Target: {target_result['error']}")
        status = " | ".join(errors)
    else:
        # Compare JSON responses
        is_matched, mismatch_detail = compare_json(
            source_result['data'],
            target_result['data']
        )
        status = mismatch_detail
    
    return {
        'source_url': source_url,
        'target_url': target_url,
        'source_size': source_size,
        'target_size': target_size,
        'status': status,
        'source_file': source_file,
        'target_file': target_file
    }


async def run_regression_tests(test_cases: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Run all regression tests asynchronously.
    
    Args:
        test_cases: DataFrame with 'SourceRequestURL' and 'TargetRequestURL' columns
    
    Returns:
        List of test results
    """
    total_requests = len(test_cases) * 2  # Source + Target
    progress = ProgressTracker(total_requests)
    
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)
    
    # Create authentication object
    auth = None
    if USERNAME and PASSWORD:
        auth = BasicAuth(USERNAME, PASSWORD)
        print(f"[OK] Using authentication for user: {USERNAME}")
    else:
        print("[WARNING] No authentication configured. Requests may fail.")
    
    # Configure session with connection pooling
    connector = aiohttp.TCPConnector(limit=MAX_CONCURRENT_REQUESTS, limit_per_host=20)
    timeout = ClientTimeout(total=REQUEST_TIMEOUT)
    
    async with ClientSession(connector=connector, timeout=timeout) as session:
        tasks = []
        
        for idx, row in test_cases.iterrows():
            source_url = row['SourceRequestURL']
            target_url = row['TargetRequestURL']
            
            task = process_test_case(
                session=session,
                semaphore=semaphore,
                index=idx,
                source_url=source_url,
                target_url=target_url,
                progress=progress,
                auth=auth
            )
            tasks.append(task)
        
        # Execute all tasks
        results = await asyncio.gather(*tasks)
    
    progress.finish()
    return results


# =====================================================
# REPORT GENERATION
# =====================================================

def generate_excel_report(results: List[Dict[str, Any]], output_file: str):
    """
    Generate formatted Excel report.
    
    Args:
        results: List of test results
        output_file: Path to output Excel file
    """
    print("\nGenerating Excel report...")
    
    # Create DataFrame
    report_data = []
    for result in results:
        report_data.append({
            'Source API URL': result['source_url'],
            'Target API URL': result['target_url'],
            'Source JSON File Size': result['source_size'],
            'Target JSON File Size': result['target_size'],
            'Status': result['status']
        })
    
    df = pd.DataFrame(report_data)
    
    # Save to Excel with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Regression Test Report', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Regression Test Report']
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 60  # Source API URL
        worksheet.column_dimensions['B'].width = 60  # Target API URL
        worksheet.column_dimensions['C'].width = 20  # Source Size
        worksheet.column_dimensions['D'].width = 20  # Target Size
        worksheet.column_dimensions['E'].width = 80  # Status
        
        # Color code status column
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row in range(2, len(df) + 2):
            status_cell = worksheet[f'E{row}']
            if status_cell.value == "Matched":
                status_cell.fill = green_fill
            else:
                status_cell.fill = red_fill
    
    print(f"[OK] Report generated: {output_file}")


# =====================================================
# MAIN EXECUTION
# =====================================================

def main():
    """Main execution function."""
    global PASSWORD
    
    print("=" * 80)
    print("API REGRESSION TESTING FRAMEWORK")
    print("=" * 80)
    
    start_time = time.time()
    
    # Load password from auth.py
    try:
        PASSWORD = get_password()
        print(f"\n[OK] Password loaded successfully")
    except Exception as e:
        print(f"\n[ERROR] Failed to load password: {e}")
        print("[WARNING] Continuing without authentication - API calls may fail")
    
    # Validate credentials
    if not USERNAME:
        print("[WARNING] USERNAME not found in .env file")
        print("[WARNING] Set USERNAME or API_USERNAME in .env file")
    
    # Step 1: Create directories
    print("\nStep 1: Setting up directories...")
    create_directories()
    
    # Step 2: Load test cases
    print(f"\nStep 2: Loading test cases from {INPUT_FILE}...")
    
    if not os.path.exists(INPUT_FILE):
        print(f"[ERROR] Input file not found: {INPUT_FILE}")
        print("Please run pl_generateTestCase.py first to generate the test cases.")
        sys.exit(1)
    
    test_cases = pd.read_excel(INPUT_FILE)
    
    required_columns = ['SourceRequestURL', 'TargetRequestURL']
    missing_columns = [col for col in required_columns if col not in test_cases.columns]
    
    if missing_columns:
        print(f"[ERROR] Missing required columns: {missing_columns}")
        sys.exit(1)
    
    # Apply test case limit if configured
    total_available = len(test_cases)
    if LIMIT_TEST_CASES is not None and LIMIT_TEST_CASES < total_available:
        test_cases = test_cases.head(LIMIT_TEST_CASES)
        print(f"[WARNING] LIMIT_TEST_CASES is active!")
        print(f"[WARNING] Processing only first {LIMIT_TEST_CASES} of {total_available} test cases")
        print(f"[WARNING] Set LIMIT_TEST_CASES = None to process all test cases")
        print()
    
    print(f"[OK] Loaded {len(test_cases)} test cases")
    print(f"[OK] Total API requests to be made: {len(test_cases) * 2}")
    
    # Step 3: Run regression tests
    print(f"\nStep 3: Executing API regression tests...")
    print(f"Configuration:")
    print(f"   - Max concurrent requests: {MAX_CONCURRENT_REQUESTS}")
    print(f"   - Request timeout: {REQUEST_TIMEOUT}s")
    print(f"   - Retry attempts: {RETRY_ATTEMPTS}")
    print()
    
    results = asyncio.run(run_regression_tests(test_cases))
    
    # Step 4: Generate report
    print(f"\nStep 4: Generating final report...")
    generate_excel_report(results, REPORT_FILE)
    
    # Step 5: Summary
    print("\n" + "=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    
    matched_count = sum(1 for r in results if r['status'] == "Matched")
    failed_count = len(results) - matched_count
    
    print(f"Total Test Cases: {len(results)}")
    print(f"Matched: {matched_count} ({matched_count/len(results)*100:.1f}%)")
    print(f"Mismatched/Failed: {failed_count} ({failed_count/len(results)*100:.1f}%)")
    print(f"\nSource API Responses: {SOURCE_RESPONSE_DIR}")
    print(f"Target API Responses: {TARGET_RESPONSE_DIR}")
    print(f"Final Report: {REPORT_FILE}")
    
    total_time = time.time() - start_time
    print(f"\nTotal execution time: {total_time:.2f} seconds")
    print("=" * 80)


if __name__ == "__main__":
    main()
